import hashlib
import base64
import requests
from openpyxl import load_workbook
import datetime
workbook1=load_workbook(('C:\\Users\\Administrator\\Desktop\\a.xlsx'))

def headers_temp():
    headers = {
        'Connection': 'keep-alive',
        'Content-Length': '984',
        'Cache-Control': 'no-cache',
        'Content-Type': 'image/x-icon',
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36',
        'Accept': 'image/webp,image/apng,image/*,*/*;q=0.8',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9',

    }
    return headers
#def url_list(mobiles,content):

if __name__ == '__main__':
    workbook1 = load_workbook(('C:\\Users\\Administrator\\Desktop\\a.xlsx'))  # 打开表，第一列是电话号码，第二列是手机号码
    url = "http://112.35.1.155:1992/sms/norsubmit"  # 接口地址
    sign = "n82SHWTQb"
    apId = "wlsbzx"
    ecName = "温岭市社会保险事业管理中心"
    addSerial = ""
    secretKey = "bd5ham"
    sheet = workbook1['Sheet1']  # 定位表单（Sheet1）
    for i in range(sheet.max_row):

        mobiles = str(sheet.cell(i + 1, 1).value)  # 手机号码
        content = sheet.cell(i + 1, 2).value  # 短信内容


        # mac:将ecName、apId、secretKey、templateId、mobiles、params、sign、addSerial按序拼接（无间隔符），通过MD5（32位小写）计算出的值。
        m = hashlib.md5() # 获取一个md5加密算法对象
        a = (ecName + apId + secretKey + mobiles + content + sign).encode("utf-8")
        # a = "政企分公司测试demo0123qwe13800138000移动改变生活。DWItALe3A".encode("utf-8")
        m.update(a)# 制定需要加密的字符串
        mac = m.hexdigest()#获取加密后的16进制字符串

        try:
            url2 = '{"content": "' + content + '", "sign": "' + sign + '", "apId": "' + apId + '", "mac": "' + mac + '", "ecName": "' + ecName + '", "addSerial": "", "secretKey": "' + secretKey + '", "mobiles": "' + mobiles + '"}'
            print(url2)
            bytes_url2 = url2.encode("utf-8")
            str_url2 = base64.b64encode(bytes_url2)  # 被编码的参数必须是二进制数据
            payload = str_url2.decode("utf-8")
            r = requests.post(url, json=payload, headers=headers_temp(), verify=False)
            run = r.json()

            # sheet.cell(i + 1, 3).value = run["msgGroup"]
            # sheet.cell(i + 1, 4).value = run["rspcod"]
            # sheet.cell(i + 1, 5).value = run["success"]
            # workbook1.save('C:\\Users\\Administrator\\Desktop\\a.xlsx')
            print(mobiles)
            print(i+1)
        except:
            print(mobiles)

    # sheet.title=datetime.datetime.now().strftime('%Y-%m-%d')
    # workbook1.save('C:\\Users\\Administrator\\Desktop\\a.xlsx')