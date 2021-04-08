import openpyxl as ol
import pandas as pd
import datetime
import win32com.client as win32
import os.path
from openpyxl.styles import Border, Side, colors
import requests
from fake_useragent import UserAgent
import re
import json
from selenium.webdriver.chrome.options import Options
import os
from PIL import Image
from io import BytesIO
from selenium import webdriver
import time
import psutil

class Get_selenium():
    def __init__(self):



        chrome_options = Options()
        chrome_options.add_argument('--headless')
        self.driver = webdriver.Chrome(chrome_options=chrome_options)
        self.cookies = []
        self.session = requests.session()
        self.headers = {}
        self.file_name= ""

    def del_xls(self):
        path = "./"
        for root, dirs, files in os.walk(path):
            for i in files:

                if i[-2:] != "py":
                    print(i)
                    os.remove(path + i)

    def parse_se(self):

        self.driver.get(
            "https://account.wps.cn/?qrcode=kdocs&logo=kdocs&cb=https%3A%2F%2Faccount.wps.cn%2Fapi%2Fv3%2Fsession%2Fcorrelate%2Fredirect%3Ft%3D1611976250971%26appid%3D375024576%26cb%3Dhttps%253A%252F%252Fwww.kdocs.cn%252FsingleSign4CST%253Fcb%253Dhttps%25253A%25252F%25252Fwww.kdocs.cn%25252Flatest%25253Ffrom%25253Ddocs")
        time.sleep(1)
        self.driver.find_element_by_id("wechat").click()
        time.sleep(1)
        self.driver.find_element_by_class_name("dialog-footer-ok").click()
        time.sleep(2)
        a = self.driver.find_element_by_xpath("//img[@id = 'miniprogramImportImg']").get_attribute("src")

        response = requests.get(a)
        # 显示二维码
        process_list = []
        for proc in psutil.process_iter():
            process_list.append(proc)
        image = Image.open(BytesIO(response.content)).show()
        time.sleep(10)
        for proc in psutil.process_iter():
            if not proc in process_list:
                proc.kill()
    def get_cookies(self):
        self.cookies =  self.driver.get_cookies()
        for cookie in self.cookies:
            self.session.cookies.set(cookie['name'], cookie['value'])
        return self.cookies
    def close(self):
        self.driver.close()


    def ret_augst(self):
        ua = UserAgent()

        return {'User-Agent':ua.random}

    def parse_downadd(self):
        self.headers = self.ret_augst()
        ret = self.session.get("https://www.kdocs.cn/latest?from=docs", headers=self.headers)
        html =ret.content.decode()
        # with open("renren.html", "w", encoding="utf-8") as f:
        #     f.write(html)
        #file_name = "2021年2月1日-2月5日.xls"
        self.file_name = input("请输入时间")
        self.file_name =self.file_name[0:4]+"年"+self.file_name[4:6]+"月"+self.file_name[6:8]+"日-"
        #print(self.file_name)
        #print(html)
        a = re.findall(r'''"fileid":"(.+?)",.*?(''' + self.file_name + ".*?.xls)", html)
        #print(a[0][0])
        #print(a[0][1])
        id_url = "https://www.kdocs.cn/3rd/drive/api/v3/groups/1040371001/files/"+a[0][0]+"/download?isblocks=false"
        parse_str = self.session.get(id_url, headers=self.headers)
        prase_con = parse_str.content.decode()
        parse_ret = json.loads(prase_con)
        return parse_ret['fileinfo']['url'],a[0][1]

    def down(self,add,file_name):
            down_url = add
            ret_huru = self.session.get(down_url, headers=self.headers)
            with open(".\\" + file_name, "wb") as f:
                f.write(ret_huru.content)
    def run(self):
        self.del_xls()
        self.parse_se()
        self.get_cookies()
        time.sleep(3)
        self.close()
        downadd,file_name = self.parse_downadd()
        self.down(downadd,file_name)

class xls_xlsx:
    def __init__(self,file_name):

        self.rootdir = os.getcwd()  # 需要转换的xls文件存放处
        self.rootdir1 = os.getcwd()  # 转换好的xlsx文件存放处
        self.files = os.listdir(self.rootdir)  # 列出xls文件夹下的所有文件
        self.num = len(self.files)  # 列出所有文件的个数
    def xlsx_ret(self):
        for i in range(self.num):  # 按文件个数执行次数
            kname = os.path.splitext(self.files[i])[1]  # 分离文件名与扩展名，返回(f_name, f_extension)元组
            if kname == '.xls':  # 判定扩展名是否为xls,屏蔽其它文件
                fname = self.rootdir  +"\\"+ self.files[i]  # 合成需要转换的路径与文件名
                fname1 = self.rootdir1 +"\\" + self.files[i]  # 合成准备存放转换好的路径与文件名
                excel = win32.gencache.EnsureDispatch('Excel.Application')  # 调用win32模块
                wb = excel.Workbooks.Open(fname)  # 打开需要转换的文件
                wb.SaveAs(fname1 + "x", FileFormat=51)  # 文件另存为xlsx扩展名的文件

                file_name = fname1+"x"

                wb.Close()
                excel.Application.Quit()
        print(file_name)
        return file_name
# for root,dirs,files in os.walk("./"):
#     file_name=""
#     for file in files:
#         if os.path.join(file) != "temp.py" :
#             file_name_1 = (os.path.join(file))
#
#             if file_name_1[-1]=="x":
#                 file_name=file_name_1
#                 print(file_name)
#file_name = "2021年2月1日-2月5日.xlsx"

class Xls_out:
    def __init__(self,file_name):
        self.df_inner = pd.read_excel(file_name, converters={"a": str, "b": str, "c": str, "d": str,"e": str})  # C:\\Users\\Administrator\\Desktop\\1.xls
    def out_ret(self):
        data1= self.df_inner.query('一 == ["大份", "小份"]')
        data1["行政服务中心人力社保窗口午餐统计表"] = range(1, len(data1) + 1)

        #print(data1)
        data1.to_excel(file_name+'1.xlsx', sheet_name='Sheet1', index=None, encoding="gbk")
        print("*"*100)
        data1= self.df_inner.query('二 == ["大份", "小份"]')
        data1["行政服务中心人力社保窗口午餐统计表"] = range(1, len(data1) + 1)

        #print(data1)
        data1.to_excel(file_name+'2.xlsx', sheet_name='Sheet1', index=None, encoding="gbk")

        data1= self.df_inner.query('三 == ["大份", "小份"]')
        data1["行政服务中心人力社保窗口午餐统计表"] = range(1, len(data1) + 1)

        #print(data1)
        data1.to_excel(file_name+'3.xlsx', sheet_name='Sheet1', index=None, encoding="gbk")

        data1= self.df_inner.query('四 == ["大份", "小份"]')
        data1["行政服务中心人力社保窗口午餐统计表"] = range(1, len(data1) + 1)

        #print(data1)
        data1.to_excel(file_name+'4.xlsx', sheet_name='Sheet1', index=None, encoding="gbk")

        data1= self.df_inner.query('五 == ["大份", "小份"]')
        data1["行政服务中心人力社保窗口午餐统计表"] = range(1, len(data1) + 1)

        #print(data1)
        data1.to_excel(file_name+'5.xlsx', sheet_name='Sheet1', index=None, encoding="gbk")
class Xls_set:
    def __init__(self,file_name):
        border_set = Border(left=Side(style='thin', color="FF000000"),
                            right=Side(style='thin', color="FF000000"),
                            top=Side(style='thin', color="FF000000"),
                            bottom=Side(style='thin', color="FF000000"))

        self.excelBook3 = ol.load_workbook(file_name)


        name3 = self.excelBook3['Sheet1']

        cell_list2 = [] # 第二行
        cell_list3 =[] # 第三行

        cell_list=name3.cell(1, 1).value # 获取第一行
        #print(cell_list)
        for i in range(1, 9):

            cell =  name3.cell(2, i).value
            if isinstance(cell, datetime.datetime): # 判断是不是日期，是的话转字符串
                time = datetime.datetime.strftime(cell, "%Y-%m-%d")
                cell_list2.append(time)

            else:
                cell_list2.append(cell)
        #print(cell_list2)
        for i in range(1, 4):
            cell_list3.append(name3.cell(3, i).value)
        #print(cell_list3)



        for a in range(1, 6):
            excelPath = ".\\"+str(a)+".xlsx"
            excelBook = ol.load_workbook(excelPath)
            name = excelBook['Sheet1']
            name.delete_rows(1,1)
            name3 = excelBook['Sheet1']

            # 插入三行
            name.insert_rows(1, 3)
            name["A1"] = cell_list
            for index, v in enumerate(cell_list2): # 获取值跟位置
                name.cell(2, index + 1).value = v
            for index, v in enumerate(cell_list3):
                name.cell(3, index + 1).value = v
            name.merge_cells("A1:C1")

            name["A52"] = cell_list
            for index, v in enumerate(cell_list2):  # 获取值跟位置
                name.cell(53, index + 1).value = v
            for index, v in enumerate(cell_list3):
                name.cell(54, index + 1).value = v
            name.merge_cells("A52:C52")

            swith = False
            for i in name.rows:
                if swith == False:
                    for j in name.values:
                        if swith == False:
                            if 617 in j:
                                swith = True
                                #print(j[0], j[1])
                                for l in range(55, 60):

                                    for m in range(1, 9):

                                        name.cell(l, m).value = (name3.cell(j[0] + 3+l-55, m).value)
                                        name.cell(j[0] + 3+l-55, m).value = None
                                break






            # 删除多余的日期
            name.delete_cols(9,14)
            if a ==1:
                name.delete_cols(5, 5 - a)
            else:
                name.delete_cols(4,a-1)
                name.delete_cols(5, 5 - a)
            for i in range(1, 6):
                name.cell(2, 1).value


            name.column_dimensions['A'].width = 10.25
            name.column_dimensions['B'].width = 12.5
            name.column_dimensions['C'].width = 8.25
            name.column_dimensions['D'].width = 8.38
            name.cell(42, 3).value = "大份"
            name.cell(43, 3).value = "小份"
            name.cell(60, 3).value = "大份"
            name.cell(61, 3).value = "小份"
            e1 = 0
            e2= 0
            e3= 0
            e4= 0
            for i in range(4, 42):
                p=(name3.cell(i,4).value)


                if p =="大份":
                    e1+=1
            name.cell(42, 4).value = e1

            for i in range(4, 42):
                p=(name3.cell(i,4).value)


                if p =="小份":
                    e2+=1
            name.cell(43, 4).value = e2

            for i in range(55, 59):
                p=(name3.cell(i,4).value)


                if p =="大份":
                    e3+=1
            name.cell(60, 4).value = e3

            for i in range(55, 59):
                p=(name3.cell(i,4).value)


                if p =="小份":
                    e4+=1
            name.cell(61, 4).value = e4




            kuang1 = 0
            kuang2 = 0
            for i in range(4, 42):
                p=(name3.cell(i,4).value)


                if p =="大份" or p == "小份":
                    kuang1+=1
            #print(kuang1)


            for i in range(55, 59):
                p=(name3.cell(i,4).value)


                if p =="大份" or p == "小份":
                    kuang2+=1
            #print(kuang2)



            kuang_list = ["a","b","c"]
            for k in kuang_list:
                for i in range(1, kuang1+4):
                    name[k+str(i)].border = border_set
                for i in range(52, kuang2+55):
                    name[k+str(i)].border = border_set

            name_xls = (name.cell(2, 4).value)

            #name_xls[0:4]

            name_xls[8:10]

            #xcelBook.save(r"C:\Users\Administrator\Downloads\7.xlsx"+)

            excelBook.save(".\\"+"人力社保窗口"+name_xls[5:7]+"月"+name_xls[8:10]+"日"+"中餐统计"+".xlsx")
            a+=1
    def delete(self):
        path = "./"
        for root, dirs, files in os.walk(path):
            for i in files:

                if i[-2:] != "py":
                    print(i)
                    os.remove(path + i)

if __name__ == '__main__':
    # get_selenium =Get_selenium()
    # get_selenium.run()
    file_name = ""
    # xls转xlsx 并获取文件名
    xls_xlsx =xls_xlsx(file_name)
    xlsx_ret = xls_xlsx.xlsx_ret()
    # 导出五张表
    Xls_out = Xls_out(xlsx_ret)
    out_ret = Xls_out.out_ret()
    # 调整表格
    Xls_set = Xls_set(xlsx_ret)