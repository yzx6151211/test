from openpyxl import load_workbook

workbook1=load_workbook(('C:\\Users\\Administrator\\Desktop\\3.xlsx'))

sheet=workbook1['Sheet1']# 定位表单（Sheet1）
for i in  range(sheet.max_row):
    column1=sheet.cell(i+1,1).value# 定位单元格（cell）.value是获取值
    column2=sheet.cell(i+1,2).value# 定位单元格（cell）.value是获取值
    print(column1)
    print(column2)



   # sheet.cell(3,2).value='妮妮'


    #workbook1.save('C:\\Users\\Administrator\\Desktop\\3.xlsx')